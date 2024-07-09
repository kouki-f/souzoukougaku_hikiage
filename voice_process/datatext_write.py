import openpyxl

wb = openpyxl.load_workbook( "question_test.xlsx" )
ws = wb[ "Sheet1" ]
f = open("yasuda/esd.list", "a", encoding="UTF-8")

line = 2
while(ws.cell( line, 3 ).value != None):
    print("処理中…："+str(line-1)+"回目")
    data = "voice_" + str(line-1) + ".wav|yasuda|JP|" + str(ws.cell( line, 3 ).value) +"\n"
    f.write(data)
    line += 1

print("処理終了")