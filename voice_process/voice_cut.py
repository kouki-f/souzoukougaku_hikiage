import openpyxl
from pydub import AudioSegment

wb = openpyxl.load_workbook( "question_test.xlsx" )
ws = wb[ "Sheet1" ]
sound = AudioSegment.from_file("00617.wav", format="wav")

line = 2
while(ws.cell( line, 1 ).value != None):
    print(ws.cell( line, 1 ).value)
    sound_start = ws.cell( line, 1 ).value * 1000
    sound_end = ws.cell( line, 2 ).value * 1000

    print("処理中…："+str(line-1)+"回目")

    sound1 = sound[sound_start:sound_end]
    filename = "raw/" + "voice_" + str(line-1) + ".wav"
    sound1.export(filename, format ="wav")
    line += 1

print("処理終了")
